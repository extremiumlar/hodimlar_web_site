"""Excel va PDF hisobotlari."""
from io import BytesIO
from datetime import datetime

from django.http import HttpResponse
from django.utils import timezone
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated

from accounts.permissions import IsHRRole
from accounts.models import User
from attendance.models import Attendance
from payroll.models import MonthlyPayroll


def _xlsx_response(filename: str):
    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def _write_xlsx(rows: list[list], headers: list[str]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2563EB")
    for row in rows:
        ws.append(row)
    for col in ws.columns:
        length = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(40, length + 2)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsHRRole])
def attendance_xlsx(request):
    date_from = request.query_params.get("date_from")
    date_to = request.query_params.get("date_to")
    qs = Attendance.objects.select_related("user").order_by("-date")
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)

    headers = [
        "Sana", "Hodim", "Kelgan", "Ketgan",
        "Kechikish (daq)", "Ishlangan (daq)", "Status", "Dam olish",
    ]
    rows = [
        [
            a.date.isoformat(),
            a.user.get_full_name() or a.user.username,
            a.check_in_time.strftime("%H:%M") if a.check_in_time else "",
            a.check_out_time.strftime("%H:%M") if a.check_out_time else "",
            a.late_minutes,
            a.worked_minutes,
            a.get_status_display(),
            "Ha" if a.is_weekend else "",
        ]
        for a in qs
    ]
    resp = _xlsx_response(f"davomat_{datetime.now():%Y%m%d_%H%M}.xlsx")
    resp.write(_write_xlsx(rows, headers))
    return resp


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsHRRole])
def payroll_xlsx(request):
    period = request.query_params.get("period", timezone.localdate().strftime("%Y-%m"))
    qs = MonthlyPayroll.objects.select_related("user").filter(period=period)
    headers = [
        "Hodim", "Davr", "Asosiy", "Dam olish qo'shimcha",
        "Bonus", "Jarima", "Kechikish jarimasi", "Yakuniy",
        "Ish kunlari", "Kechikish (daq)", "Kelmagan kun",
    ]
    rows = [
        [
            p.user.get_full_name() or p.user.username,
            p.period, p.base_salary, p.weekend_extra,
            p.bonus_total, p.penalty_total, p.late_penalty_total, p.total,
            p.worked_days, p.late_minutes, p.absent_days,
        ]
        for p in qs
    ]
    resp = _xlsx_response(f"oylik_{period}.xlsx")
    resp.write(_write_xlsx(rows, headers))
    return resp


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsHRRole])
def payroll_pdf(request, payroll_id: int):
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import mm

    p = MonthlyPayroll.objects.select_related("user").get(pk=payroll_id)
    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    w, h = A4

    c.setFillColorRGB(0.15, 0.39, 0.92)
    c.rect(0, h - 30 * mm, w, 30 * mm, fill=1, stroke=0)
    c.setFillColorRGB(1, 1, 1)
    c.setFont("Helvetica-Bold", 18)
    c.drawString(20 * mm, h - 18 * mm, "Oylik hisob-kitob")
    c.setFont("Helvetica", 10)
    c.drawString(20 * mm, h - 25 * mm, f"Davr: {p.period}")

    c.setFillColorRGB(0, 0, 0)
    y = h - 45 * mm
    c.setFont("Helvetica-Bold", 12)
    c.drawString(20 * mm, y, f"Hodim: {p.user.get_full_name() or p.user.username}")
    y -= 6 * mm
    c.setFont("Helvetica", 10)
    c.drawString(20 * mm, y, f"Bo'lim: {p.user.department.name if p.user.department else '-'}")
    y -= 12 * mm

    rows = [
        ("Asosiy oylik", p.base_salary),
        ("Dam olish kuni qo'shimchasi (+)", p.weekend_extra),
        ("Bonuslar (+)", p.bonus_total),
        ("Jarimalar (-)", p.penalty_total),
        ("Kechikish jarimasi (-)", p.late_penalty_total),
    ]
    c.setFont("Helvetica-Bold", 11)
    c.drawString(20 * mm, y, "Tarkib")
    y -= 7 * mm
    c.setFont("Helvetica", 11)
    for label, val in rows:
        c.drawString(25 * mm, y, label)
        c.drawRightString(w - 20 * mm, y, f"{val:,.2f} so'm")
        y -= 6 * mm

    y -= 4 * mm
    c.line(20 * mm, y, w - 20 * mm, y)
    y -= 8 * mm
    c.setFont("Helvetica-Bold", 14)
    c.drawString(25 * mm, y, "YAKUNIY:")
    c.drawRightString(w - 20 * mm, y, f"{p.total:,.2f} so'm")

    y -= 20 * mm
    c.setFont("Helvetica", 9)
    c.drawString(20 * mm, y, f"Ish kunlari: {p.worked_days}   Kelmagan: {p.absent_days}   "
                              f"Kechikish: {p.late_minutes} daq")

    c.showPage()
    c.save()

    resp = HttpResponse(buf.getvalue(), content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="payslip_{p.user.username}_{p.period}.pdf"'
    return resp
